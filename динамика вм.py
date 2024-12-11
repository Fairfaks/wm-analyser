import os
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart

import re


def get_yes_no_input(prompt):
    """Ensure input is either 'да' or 'нет'."""
    while True:
        answer = input(prompt).strip().lower()
        if answer in ['да', 'нет']:
            return answer == 'да'
        print("Пожалуйста, введите 'да' или 'нет'.")


def load_filter_file(file_name):
    """Load filter file and return list of items."""
    if os.path.exists(file_name):
        with open(file_name, 'r', encoding='utf-8') as f:
            return [line.strip() for line in f if line.strip()]
    return []


def apply_filters(df, urls, keywords, vitals, use_urls, use_keywords, use_vitals, analyze_only_brand):
    """Apply filters to the DataFrame."""
    # Фильтр по URL
    if use_urls and urls:
        df = df[df['Url'].isin(urls)]

    # Фильтр по ключевым словам
    if use_keywords and keywords:
        df = df[df['Query'].isin(keywords)]

    # Фильтр по витальным ключам (вхождение)
    if use_vitals and vitals:
        df = df[~df['Query'].str.contains('|'.join(map(re.escape, vitals)), case=False, na=False)]

    # Анализ только брендовых запросов
    if analyze_only_brand and vitals:
        df = df[df['Query'].str.contains('|'.join(map(re.escape, vitals)), case=False, na=False)]

    return df


def calculate_dynamics_and_color(df, value_type):
    """Calculate dynamics, prepare the data, and identify anomalies."""
    relevant_columns = [col for col in df.columns if col.endswith(f"_{value_type}")]
    result_df = df[['Query', 'Url'] + relevant_columns].copy()

    # Инициализация списков для значений
    dynamics = []
    differences = []
    percent_changes = []
    anomalies = []

    for _, row in result_df.iterrows():
        # Если данных недостаточно для анализа, заполняем None
        if len(relevant_columns) < 2:
            dynamics.append("Нет данных")
            differences.append(None)
            percent_changes.append(None)
            anomalies.append("Нет данных")
            continue

        row_start = row[relevant_columns[0]]
        row_end = row[relevant_columns[-1]]
        growth_count = 0
        decline_count = 0

        for i in range(len(relevant_columns) - 1):
            current_value = row[relevant_columns[i]]
            next_value = row[relevant_columns[i + 1]]
            if next_value > current_value:
                growth_count += 1
            elif next_value < current_value:
                decline_count += 1

        # Рассчитываем изменения
        difference = row_end - row_start if pd.notna(row_start) and pd.notna(row_end) else None
        percent_change = (
            round((difference / row_start) * 100, 2)
            if row_start and pd.notna(difference) and row_start != 0
            else None
        )
        anomalies.append(
            "Аномалия" if pd.notna(difference) and abs(difference) > 2 * row[relevant_columns].mean() else "ОК")

        if value_type in ['position', 'ctr'] and pd.notna(difference):
            difference = round(difference)

        if pd.notna(difference):
            if row_end > row_start and growth_count > decline_count:
                dynamics.append("Рост")
            elif row_end < row_start:
                dynamics.append("Падение")
            else:
                dynamics.append("Стабильно")
        else:
            dynamics.append("Нет данных")

        differences.append(difference)
        percent_changes.append(f"{percent_change}%" if percent_change is not None else None)

    # Добавляем рассчитанные данные в DataFrame
    result_df['Динамика изменений'] = dynamics
    result_df['Значение изменения'] = differences
    result_df['Процентное изменение'] = percent_changes
    result_df['Аномалия'] = anomalies

    return result_df


def add_chart_to_sheet(ws, value_type, rows_count):
    """Add a bar chart to the sheet based on daily sums."""
    # Определяем диапазон только для исходных данных (игнорируем расчетные столбцы)
    relevant_cols = [
        col for col in range(3, ws.max_column + 1)
        if not ws.cell(row=1, column=col).value in ['Динамика изменений', 'Значение изменения', 'Процентное изменение', 'Аномалия']
    ]

    # Добавляем строку для сумм в конце данных
    sum_row = rows_count
    ws.cell(row=sum_row, column=1, value="Итого")
    for col in relevant_cols:
        col_letter = get_column_letter(col)
        ws.cell(row=sum_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{rows_count - 1})")

    # Создаем гистограмму на основе сумм
    chart = BarChart()
    chart.title = f"Суммы значений по дням для {value_type.capitalize()}"
    chart.style = 13
    chart.y_axis.title = "Сумма"
    chart.x_axis.title = "Дни"

    # Добавляем данные в график
    data = Reference(ws, min_col=relevant_cols[0], max_col=relevant_cols[-1], min_row=sum_row, max_row=sum_row)
    categories = Reference(ws, min_col=relevant_cols[0], max_col=relevant_cols[-1], min_row=1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)

    # Располагаем график на листе
    ws.add_chart(chart, f"{get_column_letter(ws.max_column + 2)}2")


def process_file(input_path, output_path, urls, keywords, vitals, use_urls, use_keywords, use_vitals, analyze_only_brand):
    """Process the selected Excel file."""
    excel_data = pd.ExcelFile(input_path)
    data = excel_data.parse(excel_data.sheet_names[0])

    data = apply_filters(data, urls, keywords, vitals, use_urls, use_keywords, use_vitals, analyze_only_brand)
    wb = Workbook()

    for value_type in tqdm(['shows', 'position', 'demand', 'ctr', 'clicks'], desc="Processing metrics"):
        relevant_columns = [col for col in data.columns if col.endswith(f"_{value_type}")]
        if not relevant_columns:
            continue

        sheet_data = calculate_dynamics_and_color(data, value_type)
        ws = wb.create_sheet(title=value_type.capitalize())

        for r_idx, row in enumerate([sheet_data.columns.tolist()] + sheet_data.values.tolist(), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply coloring
        for col_idx, col_name in enumerate(relevant_columns, 3):
            for row_idx in range(2, len(sheet_data) + 2):
                current_value = ws.cell(row=row_idx, column=col_idx).value
                next_value = ws.cell(row=row_idx, column=col_idx + 1).value if col_idx + 1 < 3 + len(
                    relevant_columns) else None

                if next_value is not None:
                    if next_value > current_value:
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="00FF00", fill_type="solid")
                    elif next_value < current_value:
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FF0000", fill_type="solid")
                    else:
                        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="FFFF00", fill_type="solid")

        add_chart_to_sheet(ws, value_type, len(sheet_data) + 1)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(output_path)


# Main logic
if __name__ == "__main__":
    current_directory = os.getcwd()
    print(f"Текущая папка: {current_directory}")

    excel_files = [f for f in os.listdir(current_directory) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        print("Excel-файлы не найдены.")
    else:
        print("Найдены файлы:")
        for idx, file_name in enumerate(excel_files, start=1):
            print(f"{idx}: {file_name}")

        file_number = int(input("Укажите номер файла для обработки: ")) - 1
        if 0 <= file_number < len(excel_files):
            input_file = os.path.join(current_directory, excel_files[file_number])
            output_file = os.path.join(current_directory, f"processed_{excel_files[file_number]}")

            urls = load_filter_file('urls.txt')
            keywords = load_filter_file('keyword.txt')
            vitals = load_filter_file('vital.txt')

            use_urls = get_yes_no_input("Использовать ли фильтр по URL? Да/Нет: ")
            use_keywords = get_yes_no_input("Использовать ли фильтр по ключам? Да/Нет: ")
            use_vitals = get_yes_no_input("Удалить ли запросы с брендом? Да/Нет: ")
            analyze_only_brand = get_yes_no_input("Анализировать только запросы с брендом? Да/Нет: ")

            process_file(input_file, output_file, urls, keywords, vitals, use_urls, use_keywords, use_vitals,
                         analyze_only_brand)
            print(f"Файл сохранен: {output_file}")
        else:
            print("Некорректный номер файла.")
